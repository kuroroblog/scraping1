import openpyxl as px
import requests
from bs4 import BeautifulSoup

# リクエストするURLを格納する。
URL = "http://www.akibakei.info/plist_a.php?ar=%E9%9B%BB%E6%B0%97%E8%A1%97%E5%8D%97%E5%81%B4&pid=49"

# 取得したデータを元にExcelファイルへ書き込む。
# shopTitle : Excelへ書き込むタイトル
# shopList : お店の名前一覧
# addressList : お店の住所一覧
# itemList : 取扱商品一覧


def writeData(shopTitle, shopList, addressList, itemList):
    # Excelファイルのタイトルを更新する。
    ws.title = shopTitle
    # cell : 値
    # A1 : 店名
    # B1 : 住所
    # C1 : 取扱商品
    # cellの書き込み方 : https://pg-chain.com/python-excel-cell-write
    ws.cell(row=1, column=1, value="店名")
    ws.cell(row=1, column=2, value="住所")
    ws.cell(row=1, column=3, value="取扱商品")

    # lenについて : https://programming-study.com/technology/python-for-index/
    for idx in range(len(shopList)):
        # idxの開始位置は0から
        # cellの書き込み開始位置は2からにしたいため+2している。
        ws.cell(row=idx + 2, column=1, value=shopList[idx].text)
        ws.cell(row=idx + 2, column=2, value=addressList[idx].text)
        ws.cell(row=idx + 2, column=3, value=itemList[idx].text)

# 取得したHTMLデータを整形する。
# ws : Excelファイルを表すインスタンス
# res : HTMLデータ


def convertData(ws, res):
    shopTitle = res.find("h1").text
    shopList = res.find_all("h3")
    addressList = res.find_all("address")
    itemList = res.find_all("div", ["items"])
    writeData(shopTitle, shopList, addressList, itemList)

# URL情報からHTMLデータを取得する。


def resUrl():
    html = requests.get(URL).text
    return BeautifulSoup(html, "html.parser")


# __name__, __main__ とは? : https://note.nkmk.me/python-if-name-main/
if __name__ == '__main__':
    # Excelファイルの新規作成
    # シート操作 : https://qiita.com/taito273/items/07e4332293c2c59799d1
    book = px.Workbook()
    # 作成した新規ファイルをアクティブにする。
    ws = book.active
    convertData(ws, resUrl())

    # 書き込み処理の終了を出力する。
    print('FINISH!!')

    # 「shopInfo.xlsx」と名前をつけて、Excelファイルを保存する。
    book.save('shopInfo.xlsx')
